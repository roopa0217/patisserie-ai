"""
LangSmith evaluation for Patisserie AI's intent classifier.

Traces the REAL _classify_intent Python function (keyword rules + LLM
fallback) against the golden dataset -- not a Playground-recreated prompt --
since most queries never reach the LLM at all.

Requires LANGSMITH_API_KEY (and optionally LANGSMITH_PROJECT) in backend/.env.
Get a key from https://smith.langchain.com -> Settings -> API Keys.

Usage:
    python -m scripts.langsmith_eval
"""

from __future__ import annotations

import csv
import os
import sys
from pathlib import Path

from dotenv import load_dotenv

BACKEND_DIR = Path(__file__).resolve().parent.parent
EVALS_DIR = BACKEND_DIR.parent / "evals"
LANGSMITH_CSV = EVALS_DIR / "langsmith_dataset.csv"
WORKBOOK_PATH = EVALS_DIR / "intent_classifier_evaluation.xlsx"

load_dotenv(BACKEND_DIR / ".env")

if not (os.environ.get("LANGSMITH_API_KEY") or os.environ.get("LANGCHAIN_API_KEY")):
    print(
        "LANGSMITH_API_KEY is not set in backend/.env.\n"
        "Get a key from https://smith.langchain.com -> Settings -> API Keys,\n"
        "add it as LANGSMITH_API_KEY=... to backend/.env, then re-run this script.",
        file=sys.stderr,
    )
    sys.exit(1)

os.environ.setdefault("LANGSMITH_TRACING", "true")
os.environ.setdefault("LANGSMITH_PROJECT", "patisserie-intent-eval")

from langchain_core.messages import HumanMessage, SystemMessage  # noqa: E402
from langsmith import Client  # noqa: E402
from langsmith.evaluation import evaluate  # noqa: E402
from langsmith.schemas import Example, Run  # noqa: E402

from app.agent.graph import _classify_intent, _llm  # noqa: E402
from app.agent.prompts import INTENT_SYSTEM  # noqa: E402

DATASET_NAME = "patisserie-intent-classifier"

VALID_INTENTS = {"find_recipe", "scale_recipe", "build_indent", "check_anomaly", "general"}

JUDGE_SYSTEM = """You are evaluating a pastry-academy intent classifier.
The classifier predicts one of these categories for an instructor's message:
find_recipe, scale_recipe, build_indent, check_anomaly, general.
Given the query and the predicted category, decide if the prediction is correct.
Return JSON only: {"score": 0 or 1, "reasoning": "one sentence"}
"""


def load_rows() -> list[dict]:
    with open(LANGSMITH_CSV, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def upload_dataset(client: Client, rows: list[dict]):
    if client.has_dataset(dataset_name=DATASET_NAME):
        dataset = client.read_dataset(dataset_name=DATASET_NAME)
        print(f"Using existing dataset '{DATASET_NAME}' ({dataset.id})")
    else:
        dataset = client.create_dataset(
            dataset_name=DATASET_NAME,
            description="Patisserie AI intent classifier golden dataset (36 cases).",
        )
        client.create_examples(
            inputs=[{"query": r["query"]} for r in rows],
            outputs=[{"intent": r["intent"]} for r in rows],
            metadata=[{"id": r["id"]} for r in rows],
            dataset_id=dataset.id,
        )
        print(f"Created dataset '{DATASET_NAME}' with {len(rows)} examples ({dataset.id})")
    return dataset


def target(inputs: dict) -> dict:
    """Wraps the real classifier -- keyword rules with an LLM fallback."""
    return {"intent": _classify_intent(inputs["query"])}


def exact_match(run: Run, example: Example) -> dict:
    predicted = (run.outputs or {}).get("intent")
    reference = (example.outputs or {}).get("intent")
    return {"key": "intent_exact_match", "score": int(predicted == reference)}


def llm_judge(run: Run, example: Example) -> dict:
    """Sees only the query + predicted intent -- never the ground truth."""
    query = (example.inputs or {}).get("query", "")
    predicted = (run.outputs or {}).get("intent", "")

    llm = _llm(streaming=False)
    resp = llm.invoke(
        [
            SystemMessage(content=JUDGE_SYSTEM),
            HumanMessage(content=f"Query: {query}\nPredicted category: {predicted}"),
        ]
    )

    from scripts.eval_intent_classifier import _parse_judge_output

    score_str, reasoning = _parse_judge_output(resp.content.strip(), "score")
    try:
        score = int(score_str)
    except ValueError:
        score = 1 if score_str.strip().upper() in ("TRUE", "CORRECT", "YES") else 0

    return {"key": "intent_llm_judge", "score": score, "comment": reasoning}


def append_results_to_workbook(results) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not WORKBOOK_PATH.exists():
        print(f"{WORKBOOK_PATH} not found -- run scripts/eval_intent_classifier.py first, skipping.")
        return

    wb = load_workbook(WORKBOOK_PATH)
    # Distinct from the local "Step 5 · LLM Judge" tab (which judges whether a
    # failure matches the Step 4 category) -- this one is the LangSmith SDK's
    # general "was this prediction correct at all" judge, run via client.evaluate.
    sheet_name = "LangSmith · Judge Results"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = "LangSmith LLM-as-Judge results (sees query + predicted intent, not ground truth)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["id", "Query", "True intent", "Predicted intent", "Human Pass/Fail", "Judge score", "Judge reasoning", "Agrees w/ human?"])
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).fill = header_fill
        ws.cell(row=ws.max_row, column=col).font = header_font

    for row in results:
        judge_score = row["judge_score"]
        human_pass = row["pass_fail"] == "Pass"
        judge_says_correct = judge_score == 1
        agrees = "Yes" if judge_says_correct == human_pass else "No"
        ws.append(
            [
                row["id"], row["query"], row["intent"], row["predicted_intent"],
                row["pass_fail"], judge_score, row["judge_reasoning"], agrees,
            ]
        )

    widths = [8, 55, 16, 16, 14, 12, 55, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(WORKBOOK_PATH)
    print(f"Appended '{sheet_name}' tab to {WORKBOOK_PATH}")


def main():
    client = Client()
    rows = load_rows()
    dataset = upload_dataset(client, rows)

    print("\nRunning evaluation (Exact Match + LLM-as-Judge) against LangSmith...")
    exp_results = evaluate(
        target,
        data=DATASET_NAME,
        evaluators=[exact_match, llm_judge],
        experiment_prefix="intent-classifier",
        description="Golden-dataset eval of the real _classify_intent function.",
    )

    # Pull per-row results back out to merge into the Excel workbook.
    by_id = {r["id"]: r for r in rows}
    merged = []
    for res in exp_results:
        run = res["run"]
        example = res["example"]
        row_id = (example.metadata or {}).get("id")
        golden = by_id.get(row_id, {})
        eval_results = res.get("evaluation_results", {}).get("results", [])
        judge = next((e for e in eval_results if e.key == "intent_llm_judge"), None)
        predicted = (run.outputs or {}).get("intent", "")
        merged.append(
            {
                "id": row_id,
                "query": golden.get("query", ""),
                "intent": golden.get("intent", ""),
                "predicted_intent": predicted,
                "pass_fail": "Pass" if predicted == golden.get("intent") else "Fail",
                "judge_score": judge.score if judge else None,
                "judge_reasoning": judge.comment if judge else "",
            }
        )

    append_results_to_workbook(merged)
    project = os.environ.get("LANGSMITH_PROJECT", "default")
    print(f"\nExperiment: {exp_results.experiment_name}")
    print(f"View it at: https://smith.langchain.com -> project '{project}' -> Datasets & Experiments -> '{DATASET_NAME}'")


if __name__ == "__main__":
    main()
